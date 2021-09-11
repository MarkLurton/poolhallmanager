from datetime import datetime, timedelta
import math, time
from openpyxl import workbook, load_workbook

class PoolTable:
    def __init__(self, name, number, status, start, end, duration, cost):
        self.name = name
        self.number = number
        self.status = status
        self.start = start
        self.end = end
        self.duration = duration
        self.cost = cost

def __datetime(date_str):
    return datetime.strptime(date_str, '%B %d, %Y %H:%M')

def refresh():
    for pt in pooltables:
        wb = load_workbook('pt.xlsx')
        ws = wb.active
        if pt.status == 'Occupied':
            pt.end = datetime.now().strftime('%B %d, %Y %H:%M')
            ws[f'E{pt.number}'] = pt.end
            pt.duration = __datetime(pt.end) - __datetime(pt.start)
            if pt.duration.total_seconds() > 3600:
                ws[f'F{pt.number}'] = f'{pt.duration.total_seconds()/3600} Hours'
            else:
                ws[f'F{pt.number}'] = f'{pt.duration.total_seconds()/60} Minutes'
            pt.cost = f'${(math.ceil(pt.duration.total_seconds()/3600))*30}.00'
            ws[f'G{pt.number}'] = pt.cost
            wb.save('pt.xlsx')
    return


wb = load_workbook('pt.xlsx')
ws = wb.active

pooltable1 = PoolTable(ws['A1'].value, int(ws['B1'].value), ws['C1'].value, ws['D1'].value, ws['E1'].value, ws['F1'].value, ws['G1'].value)
pooltable2 = PoolTable(ws['A2'].value, int(ws['B2'].value), ws['C2'].value, ws['D2'].value, ws['E2'].value, ws['F2'].value, ws['G2'].value)
pooltable3 = PoolTable(ws['A3'].value, int(ws['B3'].value), ws['C3'].value, ws['D3'].value, ws['E3'].value, ws['F3'].value, ws['G3'].value)
pooltable4 = PoolTable(ws['A4'].value, int(ws['B4'].value), ws['C4'].value, ws['D4'].value, ws['E4'].value, ws['F4'].value, ws['G4'].value)
pooltable5 = PoolTable(ws['A5'].value, int(ws['B5'].value), ws['C5'].value, ws['D5'].value, ws['E5'].value, ws['F5'].value, ws['G5'].value)
pooltable6 = PoolTable(ws['A6'].value, int(ws['B6'].value), ws['C6'].value, ws['D6'].value, ws['E6'].value, ws['F6'].value, ws['G6'].value)
pooltable7 = PoolTable(ws['A7'].value, int(ws['B7'].value), ws['C7'].value, ws['D7'].value, ws['E7'].value, ws['F7'].value, ws['G7'].value)
pooltable8 = PoolTable(ws['A8'].value, int(ws['B8'].value), ws['C8'].value, ws['D8'].value, ws['E8'].value, ws['F8'].value, ws['G8'].value)
pooltable9 = PoolTable(ws['A9'].value, int(ws['B9'].value), ws['C9'].value, ws['D9'].value, ws['E9'].value, ws['F9'].value, ws['G9'].value)
pooltable10 = PoolTable(ws['A10'].value, int(ws['B10'].value), ws['C10'].value, ws['D10'].value, ws['E10'].value, ws['F10'].value, ws['G10'].value)
pooltable11 = PoolTable(ws['A11'].value, int(ws['B11'].value), ws['C11'].value, ws['D11'].value, ws['E11'].value, ws['F11'].value, ws['G11'].value)
pooltable12 = PoolTable(ws['A12'].value, int(ws['B12'].value), ws['C12'].value, ws['D12'].value, ws['E12'].value, ws['F12'].value, ws['G12'].value)

pooltables = [pooltable1, pooltable2, pooltable3, pooltable4, pooltable5, pooltable6, pooltable7, pooltable8, pooltable9, pooltable10, pooltable11, pooltable12]

while True:
    operation = input("Please enter 'open' to open a table, 'refresh' to refresh the tables, 'close' to close a table, or 'exit' to exit. ")
    if operation == 'open' or 'refresh' or 'close' or 'exit':
        break
    else:
        print(f"Sorry.'{operation} is not a valid operation.")
if operation == 'open':
    occupied = 0
    for opencheck in pooltables:
        if opencheck.status == 'Occupied':
            occupied += 1
    if occupied == 12:
        print('Sorry. All tables are currently occupied.')
    else:
        while True:
            try:
                table = int(input("Which table would you like to open? Or enter 'cancel' to cancel. "))
            except ValueError:
                print('Please input table as an integer. i.e 1, 2, 3 etc.')
                continue
            if 0 < table < 12:
                for pt in pooltables:
                    if pt.number == table:
                        if pt.status == 'Unoccupied':
                            pt.status = 'Occupied'
                            ws[f'C{pt.number}'] = pt.status
                            pt.start = datetime.now().strftime('%B %d, %Y %H:%M')
                            ws[f'D{pt.number}'] = pt.start
                            print(f'Table {table} is now occupied.')
                            wb.save('pt.xlsx')
                            refresh()
                            esc1 = True
                            break
                        else:
                            print(f'Table {table} is already occupied. Please select another table.')
                            esc1 = False
                if esc1:
                    break
            elif table == 'cancel':
                break
            else:
                print('Please input a number from 1 to 12.')
elif operation == 'refresh':
    refresh()
elif operation == 'close':
    unoccupied = 0
    for closecheck in pooltables:
        if closecheck.status == 'Unoccupied':
            unoccupied += 1
    if unoccupied == 12:
        print('Currently, no tables are occupied.')
    else:
        while True:
            try:
                table = int(input("Which table would you like to close? Or enter 'cancel' to cancel. "))
            except ValueError:
                print('Please input table as an integer. i.e 1, 2, 3 etc.')
                continue
            if 0 < table < 12:
                for pt in pooltables:
                    if pt.number == table:
                        if pt.status == 'Occupied':
                            pt.end = datetime.now().strftime('%B %d, %Y %H:%M')
                            ws[f'E{pt.number}'] = pt.end
                            pt.duration = __datetime(pt.end) - __datetime(pt.start)
                            if pt.duration.total_seconds() > 3600:
                                ws[f'F{pt.number}'] = f'{pt.duration.total_seconds()/3600} Hours'
                            else:
                                ws[f'F{pt.number}'] = f'{pt.duration.total_seconds()/60} Minutes'
                            pt.cost = f'${(math.ceil(pt.duration.total_seconds()/3600))*30}.00'
                            pt.duration = ws[f'F{pt.number}'].value
                            ws[f'G{pt.number}'] = pt.cost
                            date = datetime.now().strftime('%B-%d-%Y')
                            report = open(f"{pt.name}_{date}.txt", "w+")
                            report.write(f'Pool Table Nummber: {pt.number}\n')
                            report.write(f'Start Date and Time: {pt.start}\n')
                            report.write(f'End Date Time: {pt.end}\n')
                            report.write(f'Total Time Played: {pt.duration}\n')
                            report.write(f'Cost: {pt.cost}')
                            report.close()
                            pt.status = 'Unoccupied'
                            ws[f'C{pt.number}'] = pt.status
                            pt.start = ''
                            ws[f'D{pt.number}'] = pt.start
                            pt.end = ''
                            ws[f'E{pt.number}'] = pt.end
                            pt.duration = ''
                            ws[f'F{pt.number}'] = pt.duration
                            pt.cost = ''
                            ws[f'G{pt.number}'] = pt.cost
                            wb.save('pt.xlsx')
                            refresh()
                            esc1 = True
                            break
                        else:
                            print(f'Table {table} is not occupied. Please select the correct table to close.')
                            esc1 = False
                if esc1:
                    break
            elif table == 'cancel':
                break
            else:
                print('Please input a number from 1 to 12.')
